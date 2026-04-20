import csv
import io
import json
from datetime import date
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from django.shortcuts import get_object_or_404
from .models import PaqueteTuristico, ReservaFecha, Detalles_Venta, Agencia, Productos
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

class ExportarManifiestoAgenciaAPIView(APIView):
    """
    POST /api/agencia/gestion-logistica/exportar/
    Exporta el manifiesto de pasajeros para agencias.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        paquete_id = request.data.get('paquete_id')
        fecha = request.data.get('fecha')
        mes = request.data.get('mes') # Formato YYYY-MM
        formato = request.data.get('formato', 'XLS').upper()

        if not paquete_id or (not fecha and not mes):
            return HttpResponse(json.dumps({'error': 'Paquete y periodo (fecha o mes) son requeridos.'}), status=400, content_type='application/json')

        paquete = get_object_or_404(PaqueteTuristico, pk=paquete_id, agencia_id=request.user.pk)
        
        if mes:
            try:
                y, m = map(int, mes.split('-'))
                reservas_fecha = ReservaFecha.objects.filter(paquete=paquete, fecha__year=y, fecha__month=m).select_related('venta', 'venta__usuario')
                periodo_label = f"Mes_{mes}"
            except:
                return HttpResponse(json.dumps({'error': 'Formato de mes inválido (esperado YYYY-MM).'}), status=400, content_type='application/json')
        else:
            reservas_fecha = ReservaFecha.objects.filter(paquete=paquete, fecha=fecha).select_related('venta', 'venta__usuario')
            periodo_label = f"Fecha_{fecha}"

        if not reservas_fecha.exists():
            return HttpResponse(json.dumps({'error': f'No hay pasajeros registrados para este periodo ({mes if mes else fecha}).'}), status=404, content_type='application/json')


        turistas_data = []
        for res in reservas_fecha:
            novedades = res.venta.novedades_turistas
            if isinstance(novedades, str):
                try: novedades = json.loads(novedades)
                except: novedades = []
            
            if not isinstance(novedades, list) or not novedades:
                novedades = [{'nombres': res.venta.usuario.first_name, 'apellidos': res.venta.usuario.last_name, 'num_doc': 'N/A'}]

            for i, nov in enumerate(novedades):
                usr = res.venta.usuario
                telefono = None
                
                # Identificación y Nombre real del comprador vs pasajero
                tipo_u = 'Turista'
                id_u = nov.get('num_doc', 'N/A')
                comprador_nombre = f"{nov.get('nombres', '')} {nov.get('apellidos', '')}".strip() or usr.username

                if i == 0: 
                    # El titular puede ser una Agencia o Turista perfilado
                    if hasattr(usr, 'agencia'):
                        tipo_u = 'Agencia'
                        id_u = usr.agencia.nit or id_u
                        comprador_nombre = usr.agencia.nombre_agencia
                        telefono = usr.agencia.numero_telefonico
                    elif hasattr(usr, 'turista'):
                        tipo_u = 'Turista'
                        id_u = usr.turista.numero_identidad or id_u
                        telefono = usr.turista.numero_telefonico
                    elif hasattr(usr, 'proveedor'):
                        telefono = usr.proveedor.numero_telefonico
                
                if not telefono: telefono = nov.get('telefono') or nov.get('celular')
                
                # Búsqueda segura del detalle para evitar AttributeErrors
                det_item = res.venta.detalles_venta.filter(paquete=paquete.id).first()
                p_unit = float(det_item.precio_unitario) if det_item else 0.0
                estado_item = det_item.estado if det_item else 'Confirmado'
                fecha_v = res.venta.fecha.strftime('%Y-%m-%d %H:%M') if res.venta.fecha else 'S/F'

                turistas_data.append({
                    'fecha': fecha_v,
                    'referencia': f"TRX-{res.venta.id}",
                    'nombre': comprador_nombre,
                    'identificacion': id_u,
                    'tipo': tipo_u if i == 0 else 'Pasajero',
                    'correo': usr.email if i == 0 else 'S/R', 
                    'telefono': telefono or 'S/R',
                    'item': paquete.nombre,
                    'cantidad': 1 if i > 0 else res.cantidad,
                    'precio_unitario': p_unit if i == 0 else 0.0,
                    'total': (p_unit * res.cantidad) if i == 0 else 0.0,
                    'estado': estado_item if i == 0 else 'Participante'
                })

        filename = f"Manifiesto_{paquete.nombre.replace(' ', '_')}_{periodo_label}"
        return self._generar_reporte(turistas_data, filename, formato, paquete.nombre, mes if mes else fecha)

    def _generar_reporte(self, data, filename, formato, service_name, fecha):
        if formato == 'CSV': return ExportarManifiestoBase()._export_csv(data, filename)
        elif formato == 'XLS': return ExportarManifiestoBase()._export_xlsx(data, filename, service_name, fecha)
        elif formato == 'PDF': return ExportarManifiestoBase()._export_pdf(data, filename, service_name, fecha)
        return HttpResponse(json.dumps({'error': 'Formato no soportado.'}), status=400, content_type='application/json')

class ExportarDespachoProveedorAPIView(APIView):
    """
    POST /api/proveedor/gestion-logistica/exportar/
    Exporta el reporte de despacho para proveedores.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        producto_id = request.data.get('paquete_id') # Mantenemos el nombre del campo por compatibilidad con el front
        fecha = request.data.get('fecha')
        mes = request.data.get('mes') # Formato YYYY-MM
        formato = request.data.get('formato', 'XLS').upper()

        if not producto_id or (not fecha and not mes):
            return HttpResponse(json.dumps({'error': 'Producto y periodo (fecha o mes) son requeridos.'}), status=400, content_type='application/json')

        producto = get_object_or_404(Productos, pk=producto_id, proveedor_id=request.user.pk)
        
        if mes:
            try:
                y, m = map(int, mes.split('-'))
                detalles = Detalles_Venta.objects.filter(producto=producto_id, venta__fecha__year=y, venta__fecha__month=m).select_related('venta', 'venta__usuario')
                periodo_label = f"Mes_{mes}"
            except:
                return HttpResponse(json.dumps({'error': 'Formato de mes inválido.'}), status=400, content_type='application/json')
        else:
            detalles = Detalles_Venta.objects.filter(producto=producto_id, venta__fecha__date=fecha).select_related('venta', 'venta__usuario')
            periodo_label = f"Fecha_{fecha}"

        if not detalles.exists():
            return HttpResponse(json.dumps({'error': f'No hay pedidos para este periodo ({mes if mes else fecha}).'}), status=404, content_type='application/json')


        clientes_data = []
        for det in detalles:
            usr = det.venta.usuario
            tipo_u = 'Turista'
            id_u = 'N/A'
            nombre_u = f"{usr.first_name} {usr.last_name}".strip() or usr.username
            telefono = 'S/R'

            if hasattr(usr, 'agencia'):
                tipo_u = 'Agencia'
                id_u = usr.agencia.nit or 'N/A'
                nombre_u = usr.agencia.nombre_agencia
                telefono = usr.agencia.numero_telefonico or 'S/R'
            elif hasattr(usr, 'turista'):
                tipo_u = 'Turista'
                id_u = usr.turista.numero_identidad
                telefono = usr.turista.numero_telefonico or 'S/R'
            
            clientes_data.append({
                'nombre': nombre_u,
                'identificacion': id_u,
                'correo': usr.email,
                'telefono': telefono,
                'rol': f"Cliente ({tipo_u})",
                'referencia': f"TRX-{det.venta.id} (Cant: {det.cantidad})"
            })

        clientes_data = []
        for det in detalles:
            usr = det.venta.usuario
            tipo_u = 'Turista'
            id_u = 'N/A'
            nombre_u = f"{usr.first_name} {usr.last_name}".strip() or usr.username
            telefono = 'S/R'

            if hasattr(usr, 'agencia'):
                tipo_u = 'Agencia'
                id_u = usr.agencia.nit or 'N/A'
                nombre_u = usr.agencia.nombre_agencia
                telefono = usr.agencia.numero_telefonico or 'S/R'
            elif hasattr(usr, 'turista'):
                tipo_u = 'Turista'
                id_u = usr.turista.numero_identidad
                telefono = usr.turista.numero_telefonico or 'S/R'
            
            clientes_data.append({
                'fecha': det.venta.fecha.strftime('%Y-%m-%d %H:%M') if det.venta.fecha else 'S/F',
                'referencia': f"TRX-{det.venta.id}",
                'nombre': nombre_u,
                'identificacion': id_u,
                'tipo': tipo_u,
                'correo': usr.email,
                'telefono': telefono,
                'item': producto.nombre,
                'cantidad': det.cantidad,
                'precio_unitario': float(det.precio_unitario or 0),
                'total': float((det.precio_unitario or 0) * det.cantidad),
                'estado': det.estado
            })

        filename = f"Despacho_{producto.nombre.replace(' ', '_')}_{periodo_label}"
        return self._generar_reporte(clientes_data, filename, formato, producto.nombre, mes if mes else fecha)

    def _generar_reporte(self, data, filename, formato, service_name, fecha):
        if formato == 'CSV': return ExportarManifiestoBase()._export_csv(data, filename)
        elif formato == 'XLS': return ExportarManifiestoBase()._export_xlsx(data, filename, service_name, fecha)
        elif formato == 'PDF': return ExportarManifiestoBase()._export_pdf(data, filename, service_name, fecha)
        return HttpResponse(json.dumps({'error': 'Formato no soportado.'}), status=400, content_type='application/json')

class ExportarVentasGlobalesMensualAPIView(APIView):
    """
    POST /api/proveedor/gestion-logistica/exportar-global/
    Exporta el reporte consolidado de TODAS las ventas del proveedor para un mes.
    Contiene registros en cualquier estado.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        mes = request.data.get('mes') # Formato YYYY-MM
        formato = request.data.get('formato', 'XLS').upper()

        if not mes:
            return HttpResponse(json.dumps({'error': 'El mes es requerido para el reporte consolidado.'}), status=400, content_type='application/json')

        try:
            y, m = map(int, mes.split('-'))
        except:
            return HttpResponse(json.dumps({'error': 'Formato de mes inválido.'}), status=400, content_type='application/json')

        # 1. Obtener todos los IDs de productos del proveedor
        mis_productos_ids = Productos.objects.filter(proveedor_id=request.user.pk).values_list('id', flat=True)
        
        # 2. Obtener todos los detalles de venta de esos productos en ese mes
        detalles = Detalles_Venta.objects.filter(
            producto__in=mis_productos_ids,
            venta__fecha__year=y,
            venta__fecha__month=m
        ).select_related('venta', 'venta__usuario').order_by('-venta__fecha')

        if not detalles.exists():
            return HttpResponse(json.dumps({'error': f'No hay registros para el mes {mes}.'}), status=404, content_type='application/json')

        # 3. Mapear datos
        report_data = []
        prod_map = {p.id: p.nombre for p in Productos.objects.filter(id__in=mis_productos_ids)}

        for det in detalles:
            usr = det.venta.usuario
            tipo_u = 'Turista'
            id_u = 'N/A'
            nombre_u = f"{usr.first_name} {usr.last_name}".strip() or usr.username
            telefono = 'S/R'

            if hasattr(usr, 'agencia'):
                tipo_u = 'Agencia'
                id_u = usr.agencia.nit or 'N/A'
                nombre_u = usr.agencia.nombre_agencia
                telefono = usr.agencia.numero_telefonico or 'S/R'
            elif hasattr(usr, 'turista'):
                tipo_u = 'Turista'
                id_u = usr.turista.numero_identidad
                telefono = usr.turista.numero_telefonico or 'S/R'
            
            report_data.append({
                'fecha': det.venta.fecha.strftime('%Y-%m-%d %H:%M') if det.venta.fecha else 'S/F',
                'referencia': f"TRX-{det.venta.id}",
                'nombre': nombre_u,
                'identificacion': id_u,
                'tipo': tipo_u,
                'correo': usr.email,
                'telefono': telefono,
                'item': prod_map.get(det.producto, f'Prod#{det.producto}'),
                'cantidad': det.cantidad,
                'precio_unitario': float(det.precio_unitario or 0),
                'total': float((det.precio_unitario or 0) * det.cantidad),
                'estado': det.estado
            })

        filename = f"Reporte_Ventas_Consolidado_{mes}"
        return self._generar_reporte(report_data, filename, formato, "Consolidado Mensual", mes)

    def _generar_reporte(self, data, filename, formato, service_name, fecha):
        base = ExportarManifiestoBase()
        if formato == 'CSV': return base._export_csv(data, filename)
        elif formato == 'XLS': return base._export_xlsx(data, filename, service_name, fecha)
        elif formato == 'PDF': return base._export_pdf(data, filename, service_name, fecha)
        return HttpResponse(json.dumps({'error': 'Formato no soportado.'}), status=400, content_type='application/json')

class ExportarManifiestoBase:
    """Clase base con utilidades de exportación para no repetir lógica."""

    def _export_csv(self, data, filename):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Fecha', 'Referencia', 'Cliente', 'ID/NIT', 'Tipo', 'Correo', 'Teléfono', 'Item', 'Cant.', 'V. Unitario', 'Total', 'Estado'])
        for row in data:
            writer.writerow([
                row['fecha'], row['referencia'], row['nombre'], row['identificacion'], 
                row['tipo'], row['correo'], row['telefono'], row['item'],
                row['cantidad'], row['precio_unitario'], row['total'], row['estado']
            ])
        
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response

    def _export_xlsx(self, data, filename, tour_name, fecha):
        from openpyxl.drawing.image import Image as XLImage
        import os
        from django.conf import settings

        wb = Workbook()
        ws = wb.active
        ws.title = "Manifiesto"

        # 1. Encabezado con Logo y Nombre
        logo_path = os.path.join(settings.BASE_DIR, 'frontend_project', 'src', 'assets', 'public', 'amazon_logo.png')
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.width = 40  # Ajustar tamaño
            img.height = 40
            ws.add_image(img, 'A1')
        
        ws.merge_cells('B1:F1')
        ws['B1'] = "AMAZONIA VIVA"
        ws['B1'].font = Font(size=16, bold=True, color="10B981")
        ws['B1'].alignment = Alignment(vertical='center')
        ws.row_dimensions[1].height = 40

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid") # Emerald-600
        center_align = Alignment(horizontal='center')

        # Cabecera informativa
        ws.append([])
        ws.append(['MANIFIESTO DE PASAJEROS'])
        ws.append([f'Tour: {tour_name}'])
        ws.append([f'Fecha de Salida: {fecha}'])
        ws.append([])

        # Encabezados de tabla
        headers = ['Fecha', 'Ref.', 'Cliente', 'ID/NIT', 'Tipo', 'Item', 'Cant.', 'V. Unit', 'Total', 'Estado']
        ws.append(headers)
        
        # El índice de la fila de encabezados ahora es 7 por las filas previas
        for cell in ws[7]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Datos
        for row in data:
            ws.append([
                row['fecha'], row['referencia'], row['nombre'], row['identificacion'], 
                row['tipo'], row['item'], row['cantidad'], row['precio_unitario'], 
                row['total'], row['estado']
            ])

        # Ajustar anchos
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    def _export_pdf(self, data, filename, tour_name, fecha):
        from reportlab.lib.pagesizes import landscape
        from reportlab.platypus import Image as PDFImage
        import os
        from django.conf import settings

        buffer = io.BytesIO()
        # Usamos landscape para dar más espacio horizontal
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        
        # Estilo para celdas con wrap
        style_cell = styles['Normal']
        style_cell.fontSize = 8
        style_cell.leading = 10

        elements = []

        # 1. Encabezado Corporativo (Logo + Nombre)
        logo_path = os.path.join(settings.BASE_DIR, 'frontend_project', 'src', 'assets', 'public', 'amazon_logo.png')
        
        header_data = []
        if os.path.exists(logo_path):
            img = PDFImage(logo_path, width=40, height=40)
            header_data = [[img, Paragraph("<font size=18 color='#10B981'><b>AMAZONIA VIVA</b></font>", styles['Normal'])]]
        else:
            header_data = [[Paragraph("<font size=18 color='#10B981'><b>AMAZONIA VIVA</b></font>", styles['Normal'])]]

        header_table = Table(header_data, colWidths=[50, 400])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 20))

        # Título del reporte
        elements.append(Paragraph(f"<b>MANIFIESTO DE OPERACIONES</b>", styles['Title']))
        elements.append(Paragraph(f"<b>Tour:</b> {tour_name}", styles['Normal']))
        elements.append(Paragraph(f"<b>Fecha:</b> {fecha}", styles['Normal']))
        elements.append(Spacer(1, 15))

        # Tabla de datos
        table_data = [['Fecha', 'Ref.', 'Cliente', 'ID/NIT', 'Tipo', 'Cant.', 'Total', 'Estado']]
        for row in data:
            table_data.append([
                row['fecha'],
                row['referencia'],
                Paragraph(row['nombre'], style_cell),
                row['identificacion'],
                row['tipo'],
                row['cantidad'],
                f"${row['total']:,.2f}",
                Paragraph(row['estado'], style_cell)
            ])

        # Anchos en Landscape (Total ~732 de contenido)
        # Fecha(90) + Ref(60) + Cliente(180) + ID(80) + Tipo(70) + Cant(40) + Total(90) + Estado(120) = 730
        t = Table(table_data, colWidths=[90, 60, 180, 80, 70, 40, 90, 120], repeatRows=1)
        
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#10B981")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (2, -1), 'CENTER'), # ID y Rol centrados
            ('ALIGN', (4, 0), (5, -1), 'CENTER'), # Tel y Referencia centrada
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(t)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response
