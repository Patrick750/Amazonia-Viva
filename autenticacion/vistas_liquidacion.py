"""
Vistas para el módulo de Liquidación / Billetera Virtual.
Aplica a roles: agencia y proveedor.

Endpoints:
  GET  api/liquidacion/saldos/        → resumen de saldo
  POST api/liquidacion/solicitar-retiro/ → solicitar retiro
  GET  api/liquidacion/movimientos/   → historial paginado
  GET  api/liquidacion/exportar/      → descarga CSV
"""

import csv
import io
from decimal import Decimal

from django.http import HttpResponse
from django.db.models import Sum, Q
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from .models import Venta, Detalles_Venta, Agencia, Proveedor, Productos, PaqueteTuristico, Retiro

# ─── Tasa de comisión de la plataforma (%) ────────────────────────────────────
COMISION_PLATAFORMA = Decimal("8.00")   # 8 % sobre ventas brutas


def _get_user_item_info(user):
    """Devuelve la empresa, rol e IDs de sus productos/paquetes."""
    try:
        agencia = Agencia.objects.get(pk=user.pk)
        ids = list(PaqueteTuristico.objects.filter(agencia=agencia).values_list('id', flat=True))
        return agencia, "agencia", ids
    except Agencia.DoesNotExist:
        pass
    try:
        proveedor = Proveedor.objects.get(pk=user.pk)
        ids = list(Productos.objects.filter(proveedor=proveedor).values_list('id', flat=True))
        return proveedor, "proveedor", ids
    except Proveedor.DoesNotExist:
        pass
    return None, None, []


def _calcular_saldos(user):
    """
    Calcula los saldos reales de la billetera virtual.
    """
    empresa, rol, item_ids = _get_user_item_info(user)
    if not empresa or not item_ids:
        return {
            "saldo_total": 0.0, "saldo_disponible": 0.0, "saldo_pendiente": 0.0,
            "bruto_total": 0.0, "comision_total": 0.0, "comision_porcentaje": float(COMISION_PLATAFORMA),
            "desglose_comision": {"sobre_completado": 0.0, "sobre_pendiente": 0.0},
            "metodos_retiro": _get_metodos_default()
        }

    # Definición de estados para liquidación
    estados_disponibles = ["Entregado", "Llegó", "Realizado", "Completado"]
    estados_pendientes = ["Confirmado", "Enviado", "En Tránsito", "Pendiente de Empaque", "Procesando", "Pendiente"]

    # Filtrar ítems que pertenecen a esta empresa
    qs_items = Detalles_Venta.objects.all()
    if rol == "agencia":
        qs_items = qs_items.filter(paquete__in=item_ids)
    else:
        qs_items = qs_items.filter(producto__in=item_ids)

    # Cálculo por estado
    bruto_completado = Decimal("0")
    bruto_pendiente = Decimal("0")

    for item in qs_items:
        valor = Decimal(str(item.cantidad)) * Decimal(str(item.precio_unitario))
        if item.estado in estados_disponibles:
            bruto_completado += valor
        elif item.estado in estados_pendientes:
            bruto_pendiente += valor

    comision_completado = (bruto_completado * COMISION_PLATAFORMA / 100).quantize(Decimal("0.01"))
    comision_pendiente = (bruto_pendiente * COMISION_PLATAFORMA / 100).quantize(Decimal("0.01"))

    # Restar retiros realizados o en proceso
    retiros_activos = Retiro.objects.filter(usuario=user).exclude(estado='Rechazado').aggregate(
        total=Sum('monto')
    )['total'] or Decimal('0.00')

    neto_completado = bruto_completado - comision_completado
    # El saldo disponible es lo completado menos lo ya retirado
    saldo_disponible = neto_completado - retiros_activos
    neto_pendiente = bruto_pendiente - comision_pendiente
    saldo_total = saldo_disponible + neto_pendiente

    return {
        "saldo_total": float(saldo_total),
        "saldo_disponible": float(saldo_disponible),
        "saldo_pendiente": float(neto_pendiente),
        "retiros_realizados": float(retiros_activos),
        "bruto_total": float(bruto_completado + bruto_pendiente),
        "comision_total": float(comision_completado + comision_pendiente),
        "comision_porcentaje": float(COMISION_PLATAFORMA),
        "desglose_comision": {
            "porcentaje": float(COMISION_PLATAFORMA),
            "descripcion": "Comisión de plataforma Amazonia Viva (8%)",
            "sobre_completado": float(comision_completado),
            "sobre_pendiente": float(comision_pendiente),
            "total": float(comision_completado + comision_pendiente),
        },
        "metodos_retiro": _get_metodos_default(),
    }


def _get_metodos_default():
    return [
        {"id": "transferencia_bancaria", "nombre": "Transferencia Bancaria",
         "descripcion": "3-5 días hábiles", "icono": "bank"},
        {"id": "nequi", "nombre": "Nequi",
         "descripcion": "Inmediato", "icono": "mobile"},
        {"id": "daviplata", "nombre": "Daviplata",
         "descripcion": "Inmediato", "icono": "mobile"},
        {"id": "pse", "nombre": "PSE",
         "descripcion": "1-2 días hábiles", "icono": "pse"},
    ]


def _build_movimientos(user, page=1, page_size=15, filtro_tipo=None,
                        filtro_fecha_desde=None, filtro_fecha_hasta=None):
    """
    Construye el historial de movimientos filtrado por los ítems del usuario.
    """
    empresa, rol, item_ids = _get_user_item_info(user)
    if not empresa or not item_ids:
        return {"movimientos": [], "pagination": {"page": 1, "page_size": page_size, "total_count": 0, "total_pages": 1}}

    # Filtramos ventas que contengan ítems de este usuario
    if rol == "agencia":
        qs_items = Detalles_Venta.objects.filter(paquete__in=item_ids)
    else:
        qs_items = Detalles_Venta.objects.filter(producto__in=item_ids)

    #IDs de ventas involucradas
    venta_ids = qs_items.values_list('venta_id', flat=True).distinct()
    qs_ventas = Venta.objects.filter(id__in=venta_ids).order_by("-fecha")

    if filtro_fecha_desde:
        qs_ventas = qs_ventas.filter(fecha__date__gte=filtro_fecha_desde)
    if filtro_fecha_hasta:
        qs_ventas = qs_ventas.filter(fecha__date__lte=filtro_fecha_hasta)

    # Paginación manual inicial para poder filtrar por 'tipo' después de calcular netos
    # PERO para eficiencia, filtramos estados en BD que correspondan al tipo
    if filtro_tipo and filtro_tipo != "todos":
        items_por_tipo = qs_items.filter(estado__in=_get_estados_por_tipo(filtro_tipo))
        v_ids = items_por_tipo.values_list('venta_id', flat=True).distinct()
        qs_ventas = qs_ventas.filter(id__in=v_ids)

    total_count = qs_ventas.count()
    total_pages = max(1, (total_count + page_size - 1) // page_size)
    offset = (page - 1) * page_size
    ventas_page = qs_ventas[offset: offset + page_size]

    movimientos = []
    for v in ventas_page:
        # Calcular el total solo para los ítems de ESTA empresa en ESTA venta
        res_items = qs_items.filter(venta=v)
        if filtro_tipo and filtro_tipo != "todos":
            res_items = res_items.filter(estado__in=_get_estados_por_tipo(filtro_tipo))

        if not res_items.exists():
            continue

        bruto = sum(Decimal(str(i.cantidad)) * Decimal(str(i.precio_unitario)) for i in res_items)
        comision = (bruto * COMISION_PLATAFORMA / 100).quantize(Decimal("0.01"))
        neto = bruto - comision

        # Tomamos el estado del primer ítem como representativo del movimiento
        estado_item = res_items.first().estado
        tipo = _tipo_movimiento(estado_item)

        movimientos.append({
            "id": v.id,
            "fecha": v.fecha.isoformat(),
            "concepto": f"Venta #{v.id} — {res_items.count()} ítem(s) propios",
            "tipo": tipo,
            "estado": estado_item,
            "monto_bruto": float(bruto),
            "comision": float(comision),
            "monto_neto": float(neto),
            "referencia": f"VTA-{v.id:06d}",
        })

    return {
        "movimientos": movimientos,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total_count": total_count,
            "total_pages": total_pages,
        },
    }


def _get_estados_por_tipo(tipo):
    if tipo == "ingreso":
        return ["Entregado", "Llegó", "Realizado", "Completado"]
    if tipo == "reembolso":
        return ["Cancelado", "Reembolsado", "Devuelto", "Rechazado", "Reembolso"]
    if tipo == "pendiente":
        return ["Confirmado", "Enviado", "En Tránsito", "Pendiente de Empaque", "Procesando", "Pendiente"]
    return []

def _tipo_movimiento(estado):
    ingreso = ["Completado", "Confirmado", "Realizado", "Enviado", "Entregado", "Llegó"]
    reembolso = ["Cancelado", "Reembolsado", "Devuelto", "Rechazado"]
    if estado in ingreso:
        return "ingreso"
    if estado in reembolso:
        return "reembolso"
    return "pendiente"


def _concepto_venta(venta):
    """Genera un texto de concepto legible para la venta."""
    try:
        detalles = venta.detalles_venta.all()
        if detalles.exists():
            return f"Venta #{venta.id} — {detalles.count()} ítem(s)"
    except Exception:
        pass
    return f"Transacción #{venta.id}"


# ══════════════════════════════════════════════════════════════════════════════
# VISTAS
# ══════════════════════════════════════════════════════════════════════════════

class LiquidacionSaldosView(APIView):
    """GET /api/liquidacion/saldos/ — Resumen de saldo de la billetera."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        saldos = _calcular_saldos(request.user)
        if saldos is None:
            return Response(
                {"error": "Este endpoint es exclusivo para agencias y proveedores."},
                status=status.HTTP_403_FORBIDDEN,
            )
        return Response(saldos)


class SolicitarRetiroView(APIView):
    """POST /api/liquidacion/solicitar-retiro/ — Registrar solicitud de retiro."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        monto = request.data.get("monto")
        metodo = request.data.get("metodo")
        datos_bancarios = request.data.get("datos_bancarios", {})

        # Validaciones básicas
        if not monto or not metodo:
            return Response(
                {"error": "Los campos 'monto' y 'metodo' son requeridos."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            monto_decimal = Decimal(str(monto))
        except Exception:
            return Response(
                {"error": "Monto inválido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if monto_decimal <= 0:
            return Response(
                {"error": "El monto debe ser mayor a cero."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        saldos = _calcular_saldos(request.user)
        if saldos is None:
            return Response(
                {"error": "No autorizado para realizar retiros."},
                status=status.HTTP_403_FORBIDDEN,
            )

        disponible = Decimal(str(saldos["saldo_disponible"]))
        if monto_decimal > disponible:
            return Response(
                {
                    "error": f"Saldo insuficiente. Disponible: ${float(disponible):,.0f} COP",
                    "saldo_disponible": float(disponible),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # En una implementación real aquí se crearía un modelo SolicitudRetiro.
        # Para este MVP respondemos con confirmación optimista.
        # Guardar el retiro en la base de datos
        referencia = f"RET-{timezone.now().strftime('%Y%m%d%H%M%S')}-{request.user.pk}"
        
        retiro = Retiro.objects.create(
            usuario=request.user,
            monto=monto_decimal,
            metodo=metodo,
            datos_bancarios=datos_bancarios,
            referencia=referencia,
            estado='Pendiente'
        )

        return Response(
            {
                "mensaje": "Solicitud de retiro registrada exitosamente.",
                "referencia": referencia,
                "monto": float(monto_decimal),
                "metodo": metodo,
                "estado": retiro.estado,
                "estimado": _tiempo_estimado(metodo),
                "fecha_solicitud": retiro.fecha_solicitud.isoformat(),
            },
            status=status.HTTP_201_CREATED,
        )


def _tiempo_estimado(metodo):
    tiempos = {
        "transferencia_bancaria": "3-5 días hábiles",
        "nequi": "Inmediato (máx. 2 horas)",
        "daviplata": "Inmediato (máx. 2 horas)",
        "pse": "1-2 días hábiles",
    }
    return tiempos.get(metodo, "3-5 días hábiles")


class MovimientosView(APIView):
    """GET /api/liquidacion/movimientos/ — Historial de movimientos paginado."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        page = int(request.query_params.get("page", 1))
        page_size = int(request.query_params.get("page_size", 15))
        filtro_tipo = request.query_params.get("tipo", "todos")
        fecha_desde = request.query_params.get("fecha_desde", None)
        fecha_hasta = request.query_params.get("fecha_hasta", None)

        data = _build_movimientos(
            user=request.user,
            page=page,
            page_size=page_size,
            filtro_tipo=filtro_tipo,
            filtro_fecha_desde=fecha_desde,
            filtro_fecha_hasta=fecha_hasta,
        )
        return Response(data)


class ExportarMovimientosView(APIView):
    """GET /api/liquidacion/exportar/ — Exportar historial en formato CSV, XLS o PDF."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        filtro_tipo = request.query_params.get("tipo", "todos")
        fecha_desde = request.query_params.get("fecha_desde", None)
        fecha_hasta = request.query_params.get("fecha_hasta", None)
        formato = request.query_params.get("formato", "csv").lower()

        # Traemos todos sin paginación
        data = _build_movimientos(
            user=request.user,
            page=1,
            page_size=10000,
            filtro_tipo=filtro_tipo,
            filtro_fecha_desde=fecha_desde,
            filtro_fecha_hasta=fecha_hasta,
        )

        empresa, rol, _ = _get_user_item_info(request.user)
        nombre_empresa = "Empresa"
        if empresa:
            if rol == "agencia":
                nombre_empresa = getattr(empresa, 'nombre_agencia', 'Agencia Sin Nombre')
            else:
                nombre_empresa = getattr(empresa, 'nombre_proveedor', 'Proveedor Sin Nombre')

        if formato == "xls" or formato == "excel":
            return self._exportar_xls(data, nombre_empresa)
        elif formato == "pdf":
            return self._exportar_pdf(data, nombre_empresa)
        else:
            return self._exportar_csv(data)

    def _exportar_csv(self, data):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Referencia", "Fecha", "Concepto", "Tipo", "Estado",
            "Monto Bruto (COP)", "Comisión Plataforma (COP)", "Monto Neto (COP)"
        ])
        for m in data["movimientos"]:
            writer.writerow([
                m["referencia"], m["fecha"], m["concepto"], m["tipo"].capitalize(),
                m["estado"], m["monto_bruto"], m["comision"], m["monto_neto"]
            ])
        
        # Totales
        movs = data["movimientos"]
        writer.writerow([])
        writer.writerow(["TOTALES", "", "", "", "", 
                         sum(m['monto_bruto'] for m in movs),
                         sum(m['comision'] for m in movs),
                         sum(m['monto_neto'] for m in movs)])

        filename = f"movimientos_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response.content = b"\xef\xbb\xbf" + response.content # BOM
        return response

    def _exportar_xls(self, data, nombre_empresa):
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos Financieros"

        # Título
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Reporte de Movimientos - {nombre_empresa}"
        ws['A1'].font = Font(size=14, bold=True, color="10B981")
        ws['A1'].alignment = Alignment(horizontal='center')

        # Encabezados
        headers = ["Referencia", "Fecha", "Concepto", "Tipo", "Estado", "Bruto (COP)", "Comisión (COP)", "Neto (COP)"]
        ws.append(headers)
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for m in data["movimientos"]:
            ws.append([
                m["referencia"], m["fecha"][:10], m["concepto"], m["tipo"].capitalize(),
                m["estado"], m["monto_bruto"], m["comision"], m["monto_neto"]
            ])

        # Totales
        start_row = 3
        last_row = start_row + len(data["movimientos"])
        ws.append([])
        totals_row = last_row + 1
        ws.cell(row=totals_row, column=1, value="TOTALES").font = Font(bold=True)
        ws.cell(row=totals_row, column=6, value=sum(m['monto_bruto'] for m in data["movimientos"])).font = Font(bold=True)
        ws.cell(row=totals_row, column=7, value=sum(m['comision'] for m in data["movimientos"])).font = Font(bold=True)
        ws.cell(row=totals_row, column=8, value=sum(m['monto_neto'] for m in data["movimientos"])).font = Font(bold=True)

        # Ajuste de columnas
        for i in range(1, 9):
            ws.column_dimensions[get_column_letter(i)].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"movimientos_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _exportar_pdf(self, data, nombre_empresa):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []

        # Título
        elements.append(Paragraph(f"<font size=16 color='#10B981'><b>Amazonia Viva - Liquidación</b></font>", styles['Title']))
        elements.append(Paragraph(f"<b>Empresa:</b> {nombre_empresa}", styles['Normal']))
        elements.append(Paragraph(f"<b>Fecha de reporte:</b> {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Tabla
        table_data = [["Ref", "Fecha", "Concepto", "Tipo", "Estado", "Bruto", "Comisión", "Neto"]]
        for m in data["movimientos"]:
            table_data.append([
                m["referencia"], m["fecha"][:10], Paragraph(m["concepto"], styles['Normal']),
                m["tipo"], m["estado"], f"${m['monto_bruto']:,.0f}", f"${m['comision']:,.0f}", f"${m['monto_neto']:,.0f}"
            ])
        
        # Fila de totales
        movs = data["movimientos"]
        table_data.append(["TOTALES", "", "", "", "", 
                           f"${sum(m['monto_bruto'] for m in movs):,.0f}",
                           f"${sum(m['comision'] for m in movs):,.0f}",
                           f"${sum(m['monto_neto'] for m in movs):,.0f}"])

        table = Table(table_data, colWidths=[80, 70, 180, 60, 80, 80, 80, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#10B981")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (2, 1), (2, -1), 'LEFT'), # Concepto alineado a la izquierda
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), # Totales en negrita
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        filename = f"movimientos_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(buffer.read(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
class ListaRetirosView(APIView):
    """GET /api/liquidacion/retiros/ — Ver historial de retiros del usuario."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        retiros = Retiro.objects.filter(usuario=request.user).order_by('-fecha_solicitud')
        data = []
        for r in retiros:
            data.append({
                "id": r.id,
                "referencia": r.referencia,
                "monto": float(r.monto),
                "metodo": r.metodo,
                "estado": r.estado,
                "fecha": r.fecha_solicitud.isoformat(),
                "fecha_procesado": r.fecha_procesado.isoformat() if r.fecha_procesado else None
            })
        return Response(data)
