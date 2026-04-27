from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
import json
from .models import Productos, Proveedor
from .serializers_productos import ProductoSerializer
import csv
import io
import openpyxl
from django.http import HttpResponse
from .models import Categorias
class ProductosAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tipo_catalogo = request.query_params.get('tipo_catalogo', None)
        proveedor_id = request.query_params.get('proveedor_id', None)
        
        # Filtro base
        if proveedor_id:
            # Si se solicita explícitamente un proveedor (uso público/previsualización)
            productos = Productos.objects.filter(proveedor_id=proveedor_id)
        elif hasattr(request.user, 'proveedor'):
            # Si el usuario es proveedor y no hay ID, solo ve sus productos
            productos = Productos.objects.filter(proveedor=request.user.proveedor)
        else:
            # Para otros roles sin ID específico
            productos = Productos.objects.all()
            
        if tipo_catalogo:
            productos = productos.filter(tipo_catalogo=tipo_catalogo)
            
        serializer = ProductoSerializer(productos, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            # Asignar proveedor logueado o tomar el primero si es entorno dev local
            proveedor = None
            if hasattr(request.user, 'proveedor'):
                proveedor = request.user.proveedor
            
            if not proveedor:
                return Response({"error": "No tienes una cuenta de proveedor activa."}, status=status.HTTP_403_FORBIDDEN)

            data = {}
            for key in request.data.keys():
                if key in ['archivos_subidos', 'imagenes_eliminar']:
                    data[key] = request.data.getlist(key)
                elif key == 'caracteristicas':
                    try:
                        data[key] = json.loads(request.data.get(key))
                    except:
                        data[key] = request.data.get(key)
                elif key == 'disponible':
                    data[key] = str(request.data.get(key)).lower() == 'true'
                else:
                    data[key] = request.data.get(key)
                    
            if proveedor:
                data['proveedor'] = proveedor.id

            serializer = ProductoSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response({'mensaje': 'Producto guardado exitosamente.', 'producto': serializer.data}, status=status.HTTP_201_CREATED)
            return Response({'errores': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ProductoDetalleAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        producto = get_object_or_404(Productos, pk=pk)
        
        # Verificar propiedad
        if hasattr(request.user, 'proveedor') and producto.proveedor != request.user.proveedor:
            return Response({"error": "No tienes permiso para modificar este producto."}, status=status.HTTP_403_FORBIDDEN)
        data = {}
        for key in request.data.keys():
            if key in ['archivos_subidos', 'imagenes_eliminar']:
                data[key] = request.data.getlist(key)
            elif key == 'caracteristicas':
                try:
                    data[key] = json.loads(request.data.get(key))
                except:
                    data[key] = request.data.get(key)
            elif key == 'disponible':
                data[key] = str(request.data.get(key)).lower() == 'true'
            else:
                data[key] = request.data.get(key)
        
        serializer = ProductoSerializer(producto, data=data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({'mensaje': 'Producto actualizado con éxito.', 'producto': serializer.data}, status=status.HTTP_200_OK)
        return Response({'errores': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        producto = get_object_or_404(Productos, pk=pk)
        
        # Verificar propiedad
        if hasattr(request.user, 'proveedor') and producto.proveedor != request.user.proveedor:
            return Response({"error": "No tienes permiso para eliminar este producto."}, status=status.HTTP_403_FORBIDDEN)
            
        producto.delete()
        return Response({'mensaje': 'Producto eliminado correctamente.'}, status=status.HTTP_204_NO_CONTENT)

class CargaMasivaProductosAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Generar plantilla Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Productos"
        
        # Cabeceras
        headers = [
            "nombre", "sku", "stock", "precio", "disponible", 
            "categoria_id", "tipo_catalogo", "attr:Color", "attr:Talla", "attr:Material"
        ]
        ws.append(headers)
        
        # Fila de ejemplo
        ejemplo = [
            "Ejemplo de Producto", "SKU-001", 10, 50000, "true", 
            1, "turistas", "Rojo", "M", "Algodón"
        ]
        ws.append(ejemplo)
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="plantilla_productos.xlsx"'
        wb.save(response)
        return response

    def post(self, request):
        try:
            if not hasattr(request.user, 'proveedor'):
                return Response({"error": "No tienes una cuenta de proveedor activa."}, status=status.HTTP_403_FORBIDDEN)
            
            proveedor = request.user.proveedor
            archivo = request.FILES.get('archivo')
            
            if not archivo:
                return Response({"error": "No se ha proporcionado ningún archivo."}, status=status.HTTP_400_BAD_REQUEST)

            nombre_archivo = archivo.name.lower()
            productos_a_crear = []
            errores = []

            def procesar_fila(fila, numero_fila):
                try:
                    # Limpiar las llaves y los valores de la fila
                    cleaned_fila = {}
                    for k, v in fila.items():
                        if k:
                            key_clean = str(k).strip().lower()
                            # Eliminar prefijo attr: si lo traen, por retrocompatibilidad
                            if key_clean.startswith('attr:'):
                                key_clean = key_clean[5:].strip()
                            
                            val_clean = str(v).strip() if v is not None else ""
                            if val_clean.lower() == 'none':
                                val_clean = ""
                                
                            cleaned_fila[key_clean] = val_clean

                    nombre = cleaned_fila.get('nombre', '')
                    sku = cleaned_fila.get('sku', '')
                    stock_str = cleaned_fila.get('stock', '0')
                    precio_str = cleaned_fila.get('precio', '0')
                    disponible_str = cleaned_fila.get('disponible', 'true').lower()
                    categoria_id_str = cleaned_fila.get('categoria_id', '')
                    tipo_catalogo = cleaned_fila.get('tipo_catalogo', 'turistas').lower()

                    # Procesar atributos dinámicos (columnas extra no contempladas en base)
                    campos_base = ['nombre', 'sku', 'stock', 'precio', 'disponible', 'categoria_id', 'tipo_catalogo']
                    caracteristicas = []
                    for key, value in cleaned_fila.items():
                        if key not in campos_base and value:
                            caracteristicas.append({"clave": key.capitalize(), "valor": value})

                    # Limpieza de datos y manejo de vacíos sin arrojar error
                    try:
                        stock = int(float(stock_str)) if stock_str else 0
                    except ValueError:
                        stock = 0

                    try:
                        precio = float(precio_str) if precio_str else 0.0
                    except ValueError:
                        precio = 0.0

                    disponible = disponible_str in ['true', '1', 'si', 'yes', 's', 't']
                    
                    if tipo_catalogo not in ['turistas', 'agencias']:
                        tipo_catalogo = 'turistas'

                    # Manejar categoría de forma segura (si viene vacía o es inválida, no rompemos)
                    categoria = None
                    if categoria_id_str:
                        try:
                            cat_id = int(float(categoria_id_str))
                            categoria = Categorias.objects.filter(id=cat_id).first()
                        except ValueError:
                            pass
                    
                    if not categoria:
                        categoria = Categorias.objects.first()
                        if not categoria:
                            categoria = Categorias.objects.create(nombre="General")

                    productos_a_crear.append(Productos(
                        nombre=nombre,
                        sku=sku,
                        stock=stock,
                        precio=precio,
                        disponible=disponible,
                        categorias=categoria,
                        proveedor=proveedor,
                        tipo_catalogo=tipo_catalogo,
                        caracteristicas=caracteristicas
                    ))
                except Exception as e:
                    errores.append(f"Fila {numero_fila}: Error inesperado - {str(e)}")

            if nombre_archivo.endswith('.csv'):
                decoded_file = archivo.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string, delimiter=',')
                
                for i, row in enumerate(reader, start=2): # Start at 2 to account for header
                    if not any(str(v).strip() for v in row.values() if v is not None):
                        continue
                    procesar_fila(row, i)
                    
            elif nombre_archivo.endswith('.xlsx') or nombre_archivo.endswith('.xls'):
                wb = openpyxl.load_workbook(archivo, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                
                if not rows:
                    return Response({"error": "El archivo Excel está vacío."}, status=status.HTTP_400_BAD_REQUEST)
                    
                headers = [str(h).strip() if h else f"col_{idx}" for idx, h in enumerate(rows[0])]
                
                for i, row in enumerate(rows[1:], start=2):
                    if not any(str(v).strip() for v in row if v is not None):
                        continue
                    row_dict = dict(zip(headers, row))
                    procesar_fila(row_dict, i)
            else:
                return Response({"error": "Formato de archivo no soportado. Use .csv o .xlsx"}, status=status.HTTP_400_BAD_REQUEST)

            if productos_a_crear:
                Productos.objects.bulk_create(productos_a_crear)

            return Response({
                "mensaje": f"Se crearon {len(productos_a_crear)} productos exitosamente.",
                "errores": errores,
                "creados": len(productos_a_crear)
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': f"Error al procesar el archivo: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

